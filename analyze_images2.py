import os
from openpyxl import load_workbook

data_dir = "data"
files = os.listdir(data_dir)
excel_file = [f for f in files if f.endswith(".xlsx")][0]
excel_path = os.path.join(data_dir, excel_file)

wb = load_workbook(excel_path, data_only=True)
sheet = wb.active

with open("analyze_images2.txt", "w", encoding="utf-8") as f:
    if hasattr(sheet, '_images'):
        images = sheet._images
        for i, img in enumerate(images):
            if hasattr(img, 'anchor') and img.anchor:
                try:
                    if hasattr(img.anchor, '_from'):
                        row = img.anchor._from.row
                    else:
                        row = img.anchor.from_.row
                    
                    # Read the product code at that row (col C which is index 2, but openpyxl is 1-indexed, so col 3)
                    # Note: row from anchor is 0-indexed. So row+1 in openpyxl.
                    product_code = sheet.cell(row=row+1, column=3).value
                    f.write(f"Image {i} at row {row+1}, Code: {product_code}\n")
                except Exception as e:
                    pass
