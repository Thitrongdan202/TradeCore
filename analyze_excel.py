import os
import glob
import pandas as pd
from openpyxl import load_workbook

data_dir = "data"
files = os.listdir(data_dir)
excel_file = [f for f in files if f.endswith(".xlsx")][0]
excel_path = os.path.join(data_dir, excel_file)

with open("analyze_result.txt", "w", encoding="utf-8") as f:
    f.write(f"Found Excel file: {excel_file}\n")
    wb = load_workbook(excel_path, data_only=True)
    f.write(f"Sheets: {wb.sheetnames}\n")
    sheet = wb.active
    f.write(f"Active sheet: {sheet.title}\n")
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True):
        data.append(row)
    df = pd.DataFrame(data)
    f.write(df.to_string())

