import os
import re
import uuid
from datetime import datetime, timezone
from typing import Optional
from decimal import Decimal

from sqlalchemy.orm import Session
from openpyxl import load_workbook
import openpyxl.drawing.image

from app.core.config import get_settings
from app.models.pricing import PriceList, PriceListItem
from app.models.product import Product, ProductCategory
from app.importpipeline.report import ImportReport
from app.importpipeline.validators import RowResult, ValidationMessage, Level


def _parse_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        s = str(val).replace(",", "").strip()
        if not s:
            return None
        return Decimal(s)
    except Exception:
        return None


def import_lacasa_pricing(
    db: Session,
    file_path: str,
    original_filename: str,
    created_by_id: uuid.UUID,
    dry_run: bool = False
) -> ImportReport:
    settings = get_settings()
    report = ImportReport(entity_type="price_list", source_file=original_filename)

    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # Extract metadata
    quotation_number = None
    quotation_date = None
    effective_str = None
    vat_notes = None

    for r in range(1, 9):
        for c in range(1, 12):
            cell_val = sheet.cell(row=r, column=c).value
            if not cell_val:
                continue
            text = str(cell_val).strip()
            if "Số Báo Giá:" in text:
                quotation_number = text.split("Số Báo Giá:")[1].strip()
            elif "Ngày Báo Giá:" in text:
                quotation_date = text.split("Ngày Báo Giá:")[1].strip()
            elif "Thời gian áp dụng" in text:
                effective_str = text
            elif "giá chưa bao Gồm VAT" in text or "VAT" in text:
                vat_notes = text

    price_list = None
    if not dry_run:
        price_list = PriceList(
            name=f"Bảng giá {quotation_number or 'Mới'}",
            code=quotation_number or f"BG-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            quotation_number=quotation_number,
            quotation_date=quotation_date,
            pricing_conditions=effective_str,
            vat_notes=vat_notes,
            source_excel_file=original_filename,
            created_by_id=created_by_id,
            status="Nháp",
            is_active=True
        )
        db.add(price_list)
        db.flush()

    # Image mapping
    # anchor._from.row is 0-indexed, meaning row 67 -> Excel Row 68
    image_map = {}
    if hasattr(sheet, '_images'):
        products_img_dir = os.path.join(settings.tradecore_storage_path, "products")
        os.makedirs(products_img_dir, exist_ok=True)
        for img in sheet._images:
            if hasattr(img, 'anchor') and img.anchor:
                try:
                    if hasattr(img.anchor, '_from'):
                        row_idx = img.anchor._from.row
                    else:
                        row_idx = img.anchor.from_.row
                    
                    # Read the "MÃ HÀNG MỚI" which is in Column D (col index 3 in 0-indexed, so cell(row_idx+1, 4))
                    # Actually, let's just map image data to the row index, and then process it during row iteration.
                    image_map[row_idx] = img._data
                except Exception:
                    pass

    # Read products (headers on row 10, data starts row 11)
    current_category = None
    
    for row_idx in range(11, sheet.max_row + 1):
        stt = sheet.cell(row=row_idx, column=1).value
        nhom = sheet.cell(row=row_idx, column=2).value
        ma_hang_moi = sheet.cell(row=row_idx, column=3).value
        ma_hang_cu = sheet.cell(row=row_idx, column=4).value
        ma_xuat_hd = sheet.cell(row=row_idx, column=5).value
        qr_code = sheet.cell(row=row_idx, column=6).value
        # image is in col 7 (H) - we mapped it via anchor
        thong_tin = sheet.cell(row=row_idx, column=8).value
        gia_km = sheet.cell(row=row_idx, column=9).value

        # If it's a category header (e.g. STT is Roman numeral, or NHOM is not None but ma_hang_moi is None)
        if nhom and not ma_hang_moi and not gia_km:
            current_category = str(nhom).strip()
            continue

        if not ma_hang_moi and not ma_hang_cu:
            continue  # empty row

        primary_code = str(ma_hang_moi or ma_hang_cu).strip().upper()
        
        row_res = RowResult(row_number=row_idx, source_data={
            "MÃ HÀNG MỚI": ma_hang_moi,
            "MÃ HÀNG CŨ": ma_hang_cu,
            "NHÓM": current_category,
            "GIÁ": gia_km
        })

        if not primary_code:
            row_res.messages.append(ValidationMessage(level=Level.ERROR, code="NO_CODE", field="MÃ HÀNG MỚI", message="Thiếu mã hàng"))
            report.add_row(row_res)
            continue

        price_val = _parse_decimal(gia_km)
        if price_val is None:
            row_res.messages.append(ValidationMessage(level=Level.WARNING, code="NO_PRICE", field="GIÁ ĐẠI LÝ KM", message="Không có giá hợp lệ"))

        if not dry_run:
            # 1. Get or Create Category
            cat_id = None
            if current_category:
                cat = db.query(ProductCategory).filter(ProductCategory.name == current_category).first()
                if not cat:
                    cat = ProductCategory(name=current_category)
                    db.add(cat)
                    db.flush()
                cat_id = cat.id

            # 2. Get or Create Product
            product = db.query(Product).filter(Product.code == primary_code).first()
            if not product:
                product = Product(
                    code=primary_code,
                    name=primary_code, # Fallback, we don't have a distinct "name" column other than code and specs
                )
                db.add(product)
            
            # Update product details
            product.category_id = cat_id
            if ma_hang_cu:
                product.old_code = str(ma_hang_cu).strip()
            if ma_xuat_hd:
                product.invoice_code = str(ma_xuat_hd).strip()
            if qr_code:
                product.qr_code = str(qr_code).strip()
            if thong_tin:
                product.specifications = str(thong_tin).strip()

            # Handle image
            # openpyxl 0-indexed row = row_idx - 1
            img_data = image_map.get(row_idx - 1)
            if img_data:
                img_filename = f"{primary_code.replace('/', '_')}.png"
                img_path = os.path.join(settings.tradecore_storage_path, "products", img_filename)
                with open(img_path, "wb") as f:
                    f.write(img_data())  # img._data is a callable in some openpyxl versions? 
                    # Wait, in openpyxl, img._data is a callable returning bytes
                
                product.image_url = f"/api/v1/storage/products/{img_filename}"
            
            db.flush()

            # 3. Add to Price List
            if price_list and price_val is not None:
                item = PriceListItem(
                    price_list_id=price_list.id,
                    product_id=product.id,
                    price=float(price_val),
                    min_qty=1
                )
                db.add(item)
                
        report.add_row(row_res)

    report.finish()
    
    if not dry_run and price_list:
        report.import_run_id = str(price_list.id)
        
    return report
