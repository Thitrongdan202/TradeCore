import os
import io
import traceback
from pathlib import Path
import openpyxl
from PIL import Image

FILE_PATH = r"C:\Users\thitr\.gemini\antigravity\scratch\tradecore\data\0908. GIÁ ĐẠI LÝ_ LACASA.xlsx"
IMG_DIR = r"C:\Users\thitr\.gemini\antigravity\scratch\tradecore\backend\static\images\products"

def get_product_code_from_row(ws, row_idx):
    # Dòng tiêu đề là 9 (index 8). Cột D (index 4) là "MÃ HÀNG MỚI" (Cột 4 là 1-indexed? A=1, B=2, C=3, D=4)
    # Check lại các cột: A='', B='NO', C='NHÓM', D='MÃ HÀNG MỚI'
    # Vậy ô ở (row_idx, 4) là mã hàng mới.
    # Tuy nhiên vì nó là công thức, pandas read_excel có cache, nhưng openpyxl data_only=True có lấy được không?
    cell = ws.cell(row=row_idx, column=4)
    if cell.value:
        return str(cell.value).strip()
    return None

def extract_images():
    print(f"Extracting images from {FILE_PATH}...")
    Path(IMG_DIR).mkdir(parents=True, exist_ok=True)
    
    try:
        wb_data = openpyxl.load_workbook(FILE_PATH, data_only=True)
        ws_data = wb_data["GIÁ ĐẠI LÝ"]
        
        # openpyxl _images needs data_only=False sometimes? The images might be in the XML either way.
        wb_images = openpyxl.load_workbook(FILE_PATH, data_only=False)
        ws_images = wb_images["GIÁ ĐẠI LÝ"]
        
        images = getattr(ws_images, '_images', [])
        print(f"Found {len(images)} images.")
        
        count = 0
        for img in images:
            try:
                row_idx = None
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    row_idx = img.anchor._from.row + 1  # anchor row is 0-indexed
                
                if not row_idx:
                    continue
                
                # Lấy mã sản phẩm từ ws_data (data_only=True để lấy giá trị đã evaluate)
                product_code = get_product_code_from_row(ws_data, row_idx)
                if not product_code:
                    product_code = f"row_{row_idx}"
                    
                # Save image
                # openpyxl Image has .ref which is an openpyxl.drawing.image.Image object.
                # img._data() returns the bytes.
                img_data = img._data()
                
                # Detect extension using PIL
                pil_img = Image.open(io.BytesIO(img_data))
                ext = pil_img.format.lower()
                if ext == 'jpeg':
                    ext = 'jpg'
                    
                # We need to sanitize product code for filename
                safe_code = "".join(c if c.isalnum() else "_" for c in product_code)
                filename = f"{safe_code}.{ext}"
                filepath = os.path.join(IMG_DIR, filename)
                
                with open(filepath, "wb") as f:
                    f.write(img_data)
                
                count += 1
            except Exception as e:
                print(f"Failed to process image on row {row_idx}: {e}")
                
        print(f"Successfully extracted {count} images to {IMG_DIR}")
        
    except Exception as e:
        print(f"Error during extraction: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    extract_images()
