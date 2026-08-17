import os
import sys
import json
import warnings
from pathlib import Path
import openpyxl

# Setup Django/FastAPI path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from app.core.database import SessionLocal
from app.importpipeline.importers.products import import_products

warnings.filterwarnings("ignore", category=UserWarning)

FILE_PATH = r"C:\Users\thitr\.gemini\antigravity\scratch\tradecore\data\0908. GIÁ ĐẠI LÝ_ LACASA.xlsx"

def analyze_excel():
    print("=== EXCEL FILE ANALYSIS ===")
    print(f"File: {FILE_PATH}")
    try:
        # data_only=False to inspect formulas
        wb = openpyxl.load_workbook(FILE_PATH, data_only=False)
        print("Sheets:")
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            is_hidden = ws.sheet_state != 'visible'
            print(f"  - {sheetname} (Hidden: {is_hidden})")
            print(f"    Max Row: {ws.max_row}, Max Col: {ws.max_column}")
            
            merged = len(ws.merged_cells.ranges)
            print(f"    Merged Cells: {merged}")
            
            images = getattr(ws, '_images', [])
            image_rows = []
            for img in images:
                try:
                    # Anchor could be OneCellAnchor, TwoCellAnchor, or AbsoluteAnchor
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        row = img.anchor._from.row
                        image_rows.append(row)
                except Exception:
                    pass
            print(f"    Images: {len(images)} (Rows detected: {len(image_rows)})")
            
            # Inspect first 30 rows for formulas / external refs
            formula_count = 0
            ext_refs = set()
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), start=1):
                row_vals = []
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith('='):
                        formula_count += 1
                        if '.xlsx]' in val or '.xls]' in val:
                            ext_refs.add(val)
                    if row_idx <= 15: # Grab top 15 rows for headers
                        row_vals.append(str(val) if val else "")
                if row_idx <= 15 and any(row_vals):
                    headers.append(row_vals)

            print(f"    Formulas detected in top 30 rows: {formula_count}")
            if ext_refs:
                print(f"    External links found: {list(ext_refs)[:5]}")
            else:
                print(f"    No obvious external links in top rows.")
            
            print(f"    Sample Headers/Top rows (1-15):")
            for idx, h in enumerate(headers, 1):
                print(f"      Row {idx}: {h}")
            print()
            
    except Exception as e:
        print(f"Error loading workbook: {e}")

def dry_run_pipeline():
    print("\n=== DRY RUN IMPORT PIPELINE ===")
    db = SessionLocal()
    try:
        report = import_products(
            db=db,
            file_path=FILE_PATH,
            dry_run=True,
            skip_rows=8
        )
        print("Dry Run Report:")
        print(f"  Total rows read: {report.total_rows}")
        print(f"  Imported (valid): {report.imported_rows}")
        print(f"  Skipped: {report.skipped_rows}")
        print(f"  Errors: {report.error_rows}")
        print(f"  Warnings: {report.warning_rows}")
        
        # Sample errors
        if report.error_rows > 0:
            print("  Sample errors:")
            for r in report.row_results:
                if r.has_errors:
                    print(f"    Row {r.row_number}: {[m.message for m in r.messages if m.level.value == 'ERROR']}")
                    
        # Extracted data samples
        print("  Sample extracted valid data:")
        valid_count = 0
        for r in report.row_results:
            if not r.has_errors and valid_count < 3:
                print(f"    Row {r.row_number}: {r.mapped_data}")
                valid_count += 1
                
    except Exception as e:
        print(f"Dry run failed: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

if __name__ == "__main__":
    analyze_excel()
    dry_run_pipeline()
